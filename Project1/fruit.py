import cv2
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt
import openpyxl

#####read original image#####
img = cv2.imread('fruit blurred-noisy.tif')
cv2.imshow("(a)Original",img)

#####Gaussian Blur preprocessing#####
gaussian = cv2.GaussianBlur(img, (7,7), 0)
# cv2.imshow("Gaussian",gaussian)

#####Laplacian#####
lap = cv2.Laplacian(gaussian,cv2.CV_64F,ksize=3)
abs_lap = cv2.convertScaleAbs(lap)
cv2.imshow("(b)Laplacian",abs_lap)

#####Laplacian sharpened#####
lap_sharp = cv2.add(img,abs_lap)
cv2.imshow("(c)Laplacian Sharpened",lap_sharp)

#####Sobel gradient#####
sobel1 = cv2.Sobel(gaussian,cv2.CV_64F,1,0,ksize=3)
# cv2.imshow("(d)Sobel Gradient",sobel1)
sobel2 = cv2.Sobel(gaussian,cv2.CV_64F,0,1,ksize=3)
# cv2.imshow("(d)Sobel Gradient2",sobel2)
abs_sobel1 = cv2.convertScaleAbs(sobel1)
abs_sobel2 = cv2.convertScaleAbs(sobel2)

sobel_final = cv2.add(abs_sobel1,abs_sobel2)
cv2.imshow("(d)Sobel Gradient",sobel_final)

#####Smooth gradient#####
box = cv2.boxFilter(sobel_final,-1,(5,5),normalize=True)
cv2.imshow("(e)Smooth Gradient",box)

#####Extracted feature#####
f_img = cv2.multiply(abs_lap,box)
cv2.imshow("(f)Extracted Feature",f_img)

#####(g)=(a)+(f)#####
a_f_add = cv2.addWeighted(f_img,0.15,img,0.85,0)
cv2.imshow("(a)+(f)",a_f_add)

#####Power-law transformation#####
gamma = 0.8
final = np.array(255*(a_f_add / 255) ** gamma, dtype = 'uint8')
cv2.imshow("(h)final image",final)

#####normalization#####
norm_img = cv2.normalize(img, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
norm_abs_lap = cv2.normalize(abs_lap, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_abs_lap = (255*norm_abs_lap).astype(np.uint8)
norm_lap_sharp = cv2.normalize(lap_sharp, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_lap_sharp = (255*norm_lap_sharp).astype(np.uint8)
norm_sobel_final = cv2.normalize(sobel_final, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_sobel_final = (255*norm_sobel_final).astype(np.uint8)
norm_box = cv2.normalize(box, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_box = (255*norm_box).astype(np.uint8)
norm_f_img = cv2.normalize(f_img, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_f_img = (255*norm_f_img).astype(np.uint8)
norm_a_f_add = cv2.normalize(a_f_add, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_a_f_add = (255*norm_a_f_add).astype(np.uint8)
norm_final = cv2.normalize(final, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_final = (255*norm_final).astype(np.uint8)

#####Save required images#####
PILimage = Image.fromarray(norm_img)
PILimage.save("img/fruit_a.png",dpi=(200,200))
PILimage = Image.fromarray(norm_abs_lap)
PILimage.save("img/fruit_b.png",dpi=(200,200))
PILimage = Image.fromarray(norm_lap_sharp)
PILimage.save("img/fruit_c.png",dpi=(200,200))
PILimage = Image.fromarray(norm_sobel_final)
PILimage.save("img/fruit_d.png",dpi=(200,200))
PILimage = Image.fromarray(norm_box)
PILimage.save("img/fruit_e.png",dpi=(200,200))
PILimage = Image.fromarray(norm_f_img)
PILimage.save("img/fruit_f.png",dpi=(200,200))
PILimage = Image.fromarray(norm_a_f_add)
PILimage.save("img/fruit_g.png",dpi=(200,200))
PILimage = Image.fromarray(norm_final)
PILimage.save("img/fruit_h.png",dpi=(200,200))

#Histogram of origin
hist = cv2.calcHist([norm_img], [0], None, [256], [0, 255])
plt.plot(hist)
plt.title('Histograms')
plt.xlabel('gray level')
plt.ylabel('Number of pixels')
plt.legend(['origin','output'])
plt.show()

#Histogram of final
hist2 = cv2.calcHist([norm_final], [0], None, [256], [0, 255])
plt.plot(hist2)
plt.title('Histograms')
plt.xlabel('gray level')
plt.ylabel('Number of pixels')
plt.legend(['origin','output'])
plt.show()

#Write into excel
wb = openpyxl.load_workbook('Histograms.xlsx', data_only=True)
s1 = wb['histograms']
for i in range(0,256):
    s1.cell(i+2,4).value = int(hist[i])
    s1.cell(i+2,5).value = int(hist2[i])
wb.save('Histograms.xlsx')

cv2.waitKey()